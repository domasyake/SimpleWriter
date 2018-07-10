import openpyxl as px


class XlsWriter:
    def __init__(self, elements):
        try:
            self.__workbook = px.load_workbook('data.xlsx')
        except:
            self.__workbook = px.Workbook()
            self.sheet_init(elements)
            self.__workbook.save('data.xlsx')

        self.__workbook.save('data_bu.xlsx')

    def sheet_init(self, elements):
        sheet = self.__workbook.active
        for i in range(len(elements)):
            sheet.cell(1, i + 1).value = elements[i]

    def add_raw(self, elements):
        sheet = self.__workbook.active
        write_row = sheet.max_row + 1
        for i in range(len(elements)):
            sheet.cell(write_row, i + 1).value = elements[i]

        self.__workbook.save('data.xlsx')
        self.__workbook.save('data_bu.xlsx')
